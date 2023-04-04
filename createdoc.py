import openpyxl
import os
from docx import Document
from docx2pdf import convert

# Load the Excel workbook
workbook = openpyxl.load_workbook('/Users/stenuuesoo/Documents/autocontract/Volglased.xlsx')
sheet = workbook['Template']

# Read data from the Excel file
data = {}
for row in range(1, 19):
    tag = sheet.cell(row=row, column=1).value
    value = sheet.cell(row=row, column=2).value
    if tag is not None:
        data[tag] = value

# Load the Word document
doc = Document('/Users/stenuuesoo/Documents/autocontract/Krediidimiiidi lepingu muutmine-2.docx')

# Replace the tags with their corresponding values
for paragraph in doc.paragraphs:
    for tag, value in data.items():
        if tag is not None and value is not None:
            paragraph.text = paragraph.text.replace(tag, str(value))

# Save the modified Word document
modified_docx_path = '/Users/stenuuesoo/Documents/autocontract/{}_{}.docx'.format(data['name'].replace(' ', '_'), data['newContractDate'])
doc.save(modified_docx_path)

# Convert the modified Word document to PDF
modified_pdf_path = '/Users/stenuuesoo/Documents/autocontract/{}_{}.pdf'.format(data['name'].replace(' ', '_'), data['newContractDate'])
convert(modified_docx_path, modified_pdf_path)
